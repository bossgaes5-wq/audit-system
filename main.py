from fastapi import FastAPI, UploadFile, File, Request
from fastapi.responses import HTMLResponse
from openpyxl import load_workbook
import uuid

app = FastAPI()

users = {}
sessions = {}

@app.post("/register")
def register(email: str, password: str):
if email in users:
return {"status": "exists"}
users[email] = password
return {"status": "created"}

@app.post("/login")
def login(email: str, password: str):
if email in users and users[email] == password:
token = str(uuid.uuid4())
sessions[token] = email
return {"status": "success", "token": token}
return {"status": "error"}

@app.get("/", response_class=HTMLResponse)
def home():
return "<h2 style='text-align:center'>النظام شغال 🚀</h2>"

def read_excel(file):
wb = load_workbook(file)
ws = wb.active

```
data = []
headers = []

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        headers = [str(h).strip() for h in row]
    else:
        data.append(dict(zip(headers, row)))

return data
```

@app.post("/compare")
async def compare(request: Request, file1: UploadFile = File(...), file2: UploadFile = File(...)):
token = request.query_params.get("token")
if token not in sessions:
return {"error": "غير مصرح"}

```
data1 = read_excel(file1.file)
data2 = read_excel(file2.file)

results = []

for r1 in data1:
    net1 = (r1.get("مدين") or 0) - (r1.get("دائن") or 0)

    found = False

    for r2 in data2:
        net2 = (r2.get("مدين") or 0) - (r2.get("دائن") or 0)

        if net1 == -net2:
            found = True
            break

    if not found:
        results.append({"amount": net1, "error": "❌ خطأ"})

return results
```
